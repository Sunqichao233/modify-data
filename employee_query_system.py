#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
员工信息表查询系统
使用openpyxl模块实现Excel文件的读写和查询功能
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date
import os

class EmployeeQuerySystem:
    def __init__(self, filename="employees.xlsx"):
        self.filename = filename
        self.workbook = None
        self.worksheet = None
        
    def create_employee_table(self):
        """创建员工信息表"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "员工信息表"
        
        # 设置表头
        headers = ["员工ID", "姓名", "部门", "职位", "入职日期", "工资", "电话", "邮箱"]
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 添加示例数据
        sample_data = [
            ["E001", "张三", "技术部", "软件工程师", date(2022, 3, 15), 12000, "13800138001", "zhangsan@company.com"],
            ["E002", "李四", "销售部", "销售经理", date(2021, 8, 20), 15000, "13800138002", "lisi@company.com"],
            ["E003", "王五", "人事部", "人事专员", date(2023, 1, 10), 8000, "13800138003", "wangwu@company.com"],
            ["E004", "赵六", "技术部", "前端开发", date(2022, 11, 5), 10000, "13800138004", "zhaoliu@company.com"],
            ["E005", "钱七", "财务部", "会计", date(2020, 6, 1), 9000, "13800138005", "qianqi@company.com"],
            ["E006", "孙八", "销售部", "销售代表", date(2023, 4, 12), 7000, "13800138006", "sunba@company.com"],
            ["E007", "周九", "技术部", "数据分析师", date(2021, 12, 8), 11000, "13800138007", "zhoujiu@company.com"],
            ["E008", "吴十", "市场部", "市场专员", date(2022, 9, 25), 8500, "13800138008", "wushi@company.com"]
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                self.worksheet.cell(row=row, column=col, value=value)
        
        # 调整列宽
        column_widths = [10, 12, 12, 15, 12, 10, 15, 25]
        for col, width in enumerate(column_widths, 1):
            self.worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        self.workbook.save(self.filename)
        print(f"✅ 员工信息表已创建：{self.filename}")
    
    def load_workbook(self):
        """加载工作簿"""
        if not os.path.exists(self.filename):
            print(f"❌ 文件 {self.filename} 不存在，正在创建...")
            self.create_employee_table()
            return
        
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.worksheet = self.workbook.active
            print(f"✅ 已加载工作簿：{self.filename}")
        except Exception as e:
            print(f"❌ 加载工作簿失败：{e}")
    
    def get_all_employees(self):
        """获取所有员工信息"""
        if not self.worksheet:
            self.load_workbook()
        
        employees = []
        headers = [cell.value for cell in self.worksheet[1]]
        
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 如果员工ID不为空
                employee = dict(zip(headers, row))
                employees.append(employee)
        
        return employees
    
    def query_by_id(self, employee_id):
        """根据员工ID查询"""
        employees = self.get_all_employees()
        for emp in employees:
            if emp['员工ID'] == employee_id:
                return emp
        return None
    
    def query_by_name(self, name):
        """根据姓名查询（支持模糊查询）"""
        employees = self.get_all_employees()
        results = []
        for emp in employees:
            if name in emp['姓名']:
                results.append(emp)
        return results
    
    def query_by_department(self, department):
        """根据部门查询"""
        employees = self.get_all_employees()
        results = []
        for emp in employees:
            if emp['部门'] == department:
                results.append(emp)
        return results
    
    def query_by_salary_range(self, min_salary, max_salary):
        """根据工资范围查询"""
        employees = self.get_all_employees()
        results = []
        for emp in employees:
            salary = emp['工资']
            if isinstance(salary, (int, float)) and min_salary <= salary <= max_salary:
                results.append(emp)
        return results
    
    def query_by_join_date_range(self, start_date, end_date):
        """根据入职日期范围查询"""
        employees = self.get_all_employees()
        results = []
        for emp in employees:
            join_date = emp['入职日期']
            if isinstance(join_date, date) and start_date <= join_date <= end_date:
                results.append(emp)
        return results
    
    def add_employee(self, employee_data):
        """添加新员工"""
        if not self.worksheet:
            self.load_workbook()
        
        # 找到下一个空行
        next_row = self.worksheet.max_row + 1
        
        # 添加数据
        for col, value in enumerate(employee_data, 1):
            self.worksheet.cell(row=next_row, column=col, value=value)
        
        self.workbook.save(self.filename)
        print(f"✅ 员工 {employee_data[1]} 已添加")
    
    def update_employee(self, employee_id, field, new_value):
        """更新员工信息"""
        if not self.worksheet:
            self.load_workbook()
        
        headers = [cell.value for cell in self.worksheet[1]]
        if field not in headers:
            print(f"❌ 字段 '{field}' 不存在")
            return False
        
        field_col = headers.index(field) + 1
        
        # 查找员工行
        for row in range(2, self.worksheet.max_row + 1):
            if self.worksheet.cell(row=row, column=1).value == employee_id:
                self.worksheet.cell(row=row, column=field_col, value=new_value)
                self.workbook.save(self.filename)
                print(f"✅ 员工 {employee_id} 的 {field} 已更新为 {new_value}")
                return True
        
        print(f"❌ 未找到员工ID：{employee_id}")
        return False
    
    def delete_employee(self, employee_id):
        """删除员工"""
        if not self.worksheet:
            self.load_workbook()
        
        # 查找员工行
        for row in range(2, self.worksheet.max_row + 1):
            if self.worksheet.cell(row=row, column=1).value == employee_id:
                self.worksheet.delete_rows(row)
                self.workbook.save(self.filename)
                print(f"✅ 员工 {employee_id} 已删除")
                return True
        
        print(f"❌ 未找到员工ID：{employee_id}")
        return False
    
    def get_department_statistics(self):
        """获取部门统计信息"""
        employees = self.get_all_employees()
        dept_stats = {}
        
        for emp in employees:
            dept = emp['部门']
            if dept not in dept_stats:
                dept_stats[dept] = {'count': 0, 'total_salary': 0, 'employees': []}
            
            dept_stats[dept]['count'] += 1
            dept_stats[dept]['total_salary'] += emp['工资'] if isinstance(emp['工资'], (int, float)) else 0
            dept_stats[dept]['employees'].append(emp['姓名'])
        
        # 计算平均工资
        for dept in dept_stats:
            if dept_stats[dept]['count'] > 0:
                dept_stats[dept]['avg_salary'] = dept_stats[dept]['total_salary'] / dept_stats[dept]['count']
        
        return dept_stats
    
    def print_employee(self, employee):
        """格式化打印员工信息"""
        if not employee:
            print("❌ 未找到员工信息")
            return
        
        print("=" * 50)
        for key, value in employee.items():
            print(f"{key}: {value}")
        print("=" * 50)
    
    def print_employees(self, employees):
        """格式化打印员工列表"""
        if not employees:
            print("❌ 未找到匹配的员工")
            return
        
        print(f"\n找到 {len(employees)} 名员工：")
        print("-" * 80)
        print(f"{'员工ID':<8} {'姓名':<10} {'部门':<12} {'职位':<15} {'工资':<8}")
        print("-" * 80)
        
        for emp in employees:
            print(f"{emp['员工ID']:<8} {emp['姓名']:<10} {emp['部门']:<12} {emp['职位']:<15} {emp['工资']:<8}")
        print("-" * 80)


def main():
    """主函数 - 演示各种查询功能"""
    # 创建员工查询系统实例
    emp_system = EmployeeQuerySystem()
    
    print("🚀 员工信息表查询系统演示")
    print("=" * 60)
    
    # 1. 创建或加载员工表
    emp_system.load_workbook()
    
    # 2. 查询所有员工
    print("\n📋 所有员工信息：")
    all_employees = emp_system.get_all_employees()
    emp_system.print_employees(all_employees)
    
    # 3. 根据员工ID查询
    print("\n🔍 根据员工ID查询 (E001)：")
    employee = emp_system.query_by_id("E001")
    emp_system.print_employee(employee)
    
    # 4. 根据姓名模糊查询
    print("\n🔍 根据姓名模糊查询 (包含'张')：")
    employees = emp_system.query_by_name("张")
    emp_system.print_employees(employees)
    
    # 5. 根据部门查询
    print("\n🔍 根据部门查询 (技术部)：")
    employees = emp_system.query_by_department("技术部")
    emp_system.print_employees(employees)
    
    # 6. 根据工资范围查询
    print("\n🔍 根据工资范围查询 (10000-15000)：")
    employees = emp_system.query_by_salary_range(10000, 15000)
    emp_system.print_employees(employees)
    
    # 7. 根据入职日期范围查询
    print("\n🔍 根据入职日期范围查询 (2022年)：")
    employees = emp_system.query_by_join_date_range(date(2022, 1, 1), date(2022, 12, 31))
    emp_system.print_employees(employees)
    
    # 8. 添加新员工
    print("\n➕ 添加新员工：")
    new_employee = ["E009", "陈十一", "技术部", "测试工程师", date(2023, 10, 1), 9500, "13800138009", "chenshi@company.com"]
    emp_system.add_employee(new_employee)
    
    # 9. 更新员工信息
    print("\n✏️ 更新员工工资：")
    emp_system.update_employee("E009", "工资", 10500)
    
    # 10. 部门统计
    print("\n📊 部门统计信息：")
    dept_stats = emp_system.get_department_statistics()
    for dept, stats in dept_stats.items():
        print(f"\n{dept}:")
        print(f"  人数: {stats['count']}")
        print(f"  平均工资: {stats['avg_salary']:.2f}")
        print(f"  员工: {', '.join(stats['employees'])}")
    
    # 11. 交互式查询菜单
    interactive_menu(emp_system)

def interactive_menu(emp_system):
    """交互式查询菜单"""
    while True:
        print("\n" + "=" * 60)
        print("📋 员工信息查询菜单")
        print("=" * 60)
        print("1. 根据员工ID查询")
        print("2. 根据姓名查询")
        print("3. 根据部门查询")
        print("4. 根据工资范围查询")
        print("5. 查看所有员工")
        print("6. 添加员工")
        print("7. 更新员工信息")
        print("8. 删除员工")
        print("9. 部门统计")
        print("0. 退出")
        print("=" * 60)
        
        choice = input("请选择操作 (0-9): ").strip()
        
        if choice == "1":
            emp_id = input("请输入员工ID: ").strip()
            employee = emp_system.query_by_id(emp_id)
            emp_system.print_employee(employee)
            
        elif choice == "2":
            name = input("请输入姓名（支持模糊查询）: ").strip()
            employees = emp_system.query_by_name(name)
            emp_system.print_employees(employees)
            
        elif choice == "3":
            dept = input("请输入部门名称: ").strip()
            employees = emp_system.query_by_department(dept)
            emp_system.print_employees(employees)
            
        elif choice == "4":
            try:
                min_sal = float(input("请输入最低工资: "))
                max_sal = float(input("请输入最高工资: "))
                employees = emp_system.query_by_salary_range(min_sal, max_sal)
                emp_system.print_employees(employees)
            except ValueError:
                print("❌ 请输入有效的数字")
                
        elif choice == "5":
            employees = emp_system.get_all_employees()
            emp_system.print_employees(employees)
            
        elif choice == "6":
            print("请输入新员工信息：")
            try:
                emp_id = input("员工ID: ").strip()
                name = input("姓名: ").strip()
                dept = input("部门: ").strip()
                position = input("职位: ").strip()
                join_date_str = input("入职日期 (YYYY-MM-DD): ").strip()
                join_date = datetime.strptime(join_date_str, "%Y-%m-%d").date()
                salary = float(input("工资: "))
                phone = input("电话: ").strip()
                email = input("邮箱: ").strip()
                
                new_emp = [emp_id, name, dept, position, join_date, salary, phone, email]
                emp_system.add_employee(new_emp)
            except Exception as e:
                print(f"❌ 添加失败：{e}")
                
        elif choice == "7":
            emp_id = input("请输入要更新的员工ID: ").strip()
            field = input("请输入要更新的字段名: ").strip()
            new_value = input("请输入新值: ").strip()
            
            # 尝试转换数值类型
            if field == "工资":
                try:
                    new_value = float(new_value)
                except ValueError:
                    print("❌ 工资必须是数字")
                    continue
            elif field == "入职日期":
                try:
                    new_value = datetime.strptime(new_value, "%Y-%m-%d").date()
                except ValueError:
                    print("❌ 日期格式错误，请使用 YYYY-MM-DD")
                    continue
            
            emp_system.update_employee(emp_id, field, new_value)
            
        elif choice == "8":
            emp_id = input("请输入要删除的员工ID: ").strip()
            confirm = input(f"确认删除员工 {emp_id}? (y/N): ").strip().lower()
            if confirm == 'y':
                emp_system.delete_employee(emp_id)
            else:
                print("❌ 取消删除")
                
        elif choice == "9":
            dept_stats = emp_system.get_department_statistics()
            print("\n📊 部门统计信息：")
            print("-" * 60)
            for dept, stats in dept_stats.items():
                print(f"{dept}: {stats['count']}人, 平均工资: {stats['avg_salary']:.2f}")
            print("-" * 60)
            
        elif choice == "0":
            print("👋 感谢使用员工信息查询系统！")
            break
            
        else:
            print("❌ 无效选择，请重新输入")

if __name__ == "__main__":
    main()